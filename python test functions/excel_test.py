import openpyxl

file_name='N:/PROJECTS/MRISimProject/autoMRISimQA/dailyQAExcelReport/DailyQAResultsLaser.xlsx'


def readExceQAResults(file_name):
    
    '''
    Read Excel QA results to update the QA tracker.
    '''

    wb= openpyxl.load_workbook(file_name)

    sheet=wb.get_sheet_by_name('Sheet1') 

    max_row=sheet.get_highest_row() 
    
    

    SNR=sheet.cell(row=max_row,column=2).value
    
    Uniformity=sheet.cell(row=max_row,column=3).value
    
    Contrast=sheet.cell(row=max_row,column=4).value
    
    Ghosting=sheet.cell(row=max_row,column=5).value 
    
    
    D45=sheet.cell(row=max_row,column=6).value
    
    D135=sheet.cell(row=max_row,column=7).value
    
    Output=sheet.cell(row=max_row,column=8).value
    
    LaserX=sheet.cell(row=max_row,column=9).value
    
    LaserY=sheet.cell(row=max_row,column=10).value
    
    LaserZ=sheet.cell(row=max_row,column=11).value
    
    return (SNR,Uniformity,Contrast,Ghosting,D45,D135,Output,LaserX,LaserY,LaserZ)
    
    
if __name__=='__main__':
    
    
    test=readExceQAResults(file_name)
    
    print test
    